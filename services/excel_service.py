import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models.ont_model import ONTCollection

class ExcelService:
    """Servicio para generar reportes en Excel"""
    
    @staticmethod
    def generar_reporte(ont_collection: ONTCollection) -> io.BytesIO:
        """Genera un archivo Excel con los datos de las ONTs"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte ONTs"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            "ID ONU", "DESCRIPCION", "ONT RX", "OLT RX", "DIFERENCIA",
            "TEMPERATURA", "DISTANCIA", "ESTADO", "HORA CAIDA", "ULTIMA CAIDA"
        ]
        
        # Aplicar encabezados con estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Agregar datos
        for row, ont in enumerate(ont_collection.onts, 2):
            ws.cell(row=row, column=1, value=ont.id)
            ws.cell(row=row, column=2, value=ont.descripcion)
            ws.cell(row=row, column=3, value=ont.ont_rx)
            ws.cell(row=row, column=4, value=ont.olt_rx)
            
            # Diferencia con color condicional
            diff_cell = ws.cell(row=row, column=5, value=ont.rx_diff)
            if ont.has_critical_rx_diff():
                diff_cell.font = Font(color="FF0000")  # Rojo para valores críticos
            
            ws.cell(row=row, column=6, value=ont.temperature)
            ws.cell(row=row, column=7, value=ont.distance)
            
            # Estado con color
            estado_cell = ws.cell(row=row, column=8, value=ont.estado)
            if ont.is_online():
                estado_cell.font = Font(color="00AA00")  # Verde para online
            else:
                estado_cell.font = Font(color="FF0000")  # Rojo para offline
            
            ws.cell(row=row, column=9, value=ont.last_down_time)
            ws.cell(row=row, column=10, value=ont.last_down_cause)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream
    